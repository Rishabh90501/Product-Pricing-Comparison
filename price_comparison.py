"""
Price Comparison Tool
=====================
Scrapes product Name, Price, and Rating for a laptop model (default: "HP Pavilion")
from Flipkart and Amazon.in, compares them, and exports:
  1. A comparison report  -> comparison_report.csv  (+ comparison_report.xlsx)
  2. A price bar chart    -> price_comparison_chart.png

Approach:
  - Fast path: requests + BeautifulSoup (no browser needed) with realistic,
    rotating browser headers so the sites serve normal HTML.
  - Bot-check fallback: if a site serves a CAPTCHA / "are you a robot" page
    (or returns no product cards), the tool automatically re-loads the page in
    a real headless Chrome browser via Selenium, which passes the check.
  - Defensive parsing: every field is wrapped in try/except so one missing
    field or layout change never crashes the run.
  - Respectful scraping: max 10 items per site (<50 total), polite delay
    between requests.

Usage:
    # Interactive - just run it and type the model when prompted:
    python price_comparison.py

    # Or pass the model directly on the command line (quotes optional):
    python price_comparison.py Dell Inspiron 15
    python price_comparison.py "HP Pavilion"

    # Optionally choose how many items per site (auto-capped for polite scraping):
    python price_comparison.py "Lenovo IdeaPad" --max-items 8
"""

import os
import re
import sys
import time
import shutil
import random
import argparse

import requests
import numpy as np
import pandas as pd
import matplotlib

matplotlib.use("Agg")  # render charts without a display
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup

# ----------------------------- configuration ------------------------------ #

DEFAULT_QUERY = "HP Pavilion laptop"  # used when no product is given on the CLI/prompt
MAX_ITEMS_PER_SITE = 10          # default items per site (override with --max-items)
MAX_TOTAL_ITEMS = 48             # hard cap across all sites (site usage policy: <50)
REQUEST_TIMEOUT = 20             # seconds
POLITE_DELAY = (1.5, 3.0)        # random delay range between requests (s)
HEADLESS = True                  # set False to watch the Selenium browser work
PRODUCTS_DIR = "Products"        # per-product report folders live under here
IMAGES_SUBDIR = "images"         # charts go in Products/<ProductName>/images/

# Brand-ish colors so every chart is consistent (Amazon navy, Flipkart yellow).
COLORS = {"Flipkart": "#f8c200", "Amazon": "#232f3e"}

# A few real desktop User-Agents; one is picked at random per request so a
# single UA fingerprint is less likely to get flagged as a bot.
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 "
    "(KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) "
    "Gecko/20100101 Firefox/127.0",
]

HEADERS = {
    "User-Agent": USER_AGENTS[0],  # overridden per-request by _pick_headers()
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,*/*;q=0.8"
    ),
    "Accept-Language": "en-IN,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

# Substrings that reveal a bot-check / CAPTCHA interstitial instead of results.
BOT_CHECK_SIGNS = (
    "api-services-support@amazon",
    "to discuss automated access",
    "enter the characters you see below",
    "type the characters you see in this image",
    "/errors/validatecaptcha",
    "robot check",
    "are you a human",
    "not a robot",
    "captcha",
)

PRICE_RE = re.compile(r"₹\s*([\d,]+)")

_DRIVER = None  # lazily-created shared Selenium browser (see _get_driver)


# ------------------------------- helpers ---------------------------------- #

def polite_sleep():
    """Small random pause so we do not hammer the servers."""
    time.sleep(random.uniform(*POLITE_DELAY))


def clean_price(text):
    """'₹58,490' -> 58490 (int). Returns None when no price found."""
    if not text:
        return None
    match = PRICE_RE.search(text)
    if match:
        return int(match.group(1).replace(",", ""))
    # fallback: any digit group like 58,490 or 58490
    match = re.search(r"([\d]{2,3}(?:,\d{3})+|\d{4,7})", text)
    return int(match.group(1).replace(",", "")) if match else None


def clean_rating(text):
    """'4.3 out of 5 stars' or '4.3' -> 4.3 (float). None when missing."""
    if not text:
        return None
    match = re.search(r"(\d\.?\d?)\s*(?:out of|$|\s)", text.strip())
    try:
        rating = float(match.group(1)) if match else float(text.strip())
        return rating if 0 <= rating <= 5 else None
    except (ValueError, AttributeError):
        return None


def _pick_headers():
    """Copy of HEADERS with a randomly chosen User-Agent for this request."""
    headers = dict(HEADERS)
    headers["User-Agent"] = random.choice(USER_AGENTS)
    return headers


def looks_blocked(html):
    """True if the HTML looks like a bot-check / CAPTCHA page, not results."""
    low = html.lower()
    return any(sign in low for sign in BOT_CHECK_SIGNS)


def _fetch_requests(url, session):
    """Fast path. Returns (soup, html) or (None, '') on network/HTTP error."""
    try:
        resp = session.get(url, headers=_pick_headers(), timeout=REQUEST_TIMEOUT)
        if resp.status_code != 200:
            print(f"    [!] HTTP {resp.status_code} via requests")
            return None, ""
        return BeautifulSoup(resp.text, "lxml"), resp.text
    except requests.RequestException as exc:
        print(f"    [!] Request failed: {exc}")
        return None, ""


def _get_driver():
    """Create (once) and return a stealth headless Chrome driver.

    Returns None if Selenium or Chrome is unavailable, so the caller can
    degrade gracefully instead of crashing. Selenium 4 auto-downloads the
    matching chromedriver via Selenium Manager - no manual driver needed.
    """
    global _DRIVER
    if _DRIVER is not None:
        return _DRIVER
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
    except ImportError:
        print("    [!] Selenium not installed - browser fallback unavailable. "
              "Install with: pip install selenium")
        return None
    try:
        opts = Options()
        if HEADLESS:
            opts.add_argument("--headless=new")
        # Flags that make headless Chrome look like an ordinary user browser.
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--window-size=1400,900")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--lang=en-IN")
        opts.add_argument(f"--user-agent={random.choice(USER_AGENTS)}")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        driver = webdriver.Chrome(options=opts)
        # Hide the navigator.webdriver flag that anti-bot JS checks for.
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator,'webdriver',"
                       "{get: () => undefined})"},
        )
        _DRIVER = driver
        return _DRIVER
    except Exception as exc:
        print(f"    [!] Could not start Chrome for Selenium: {exc}")
        return None


def _fetch_selenium(url, card_selector):
    """Load the page in a real browser (defeats bot-check). Returns soup/None."""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException

    driver = _get_driver()
    if driver is None:
        return None

    for attempt in range(2):
        try:
            driver.get(url)
            # Wait for real product cards to appear (JS-rendered content).
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, card_selector))
                )
            except TimeoutException:
                pass  # take whatever loaded; may still be usable or blocked
            html = driver.page_source
            soup = BeautifulSoup(html, "lxml")
            if soup.select(card_selector):
                return soup
            if looks_blocked(html) and attempt == 0:
                print("    [i] Browser still shows a check page; retrying once...")
                time.sleep(random.uniform(3, 5))
                continue
            return soup
        except Exception as exc:
            print(f"    [!] Selenium error: {exc}")
            return None
    return None


def close_driver():
    """Shut down the shared browser if one was started."""
    global _DRIVER
    if _DRIVER is not None:
        try:
            _DRIVER.quit()
        except Exception:
            pass
        _DRIVER = None


def get_page(url, card_selector, site_name):
    """Return a BeautifulSoup page that contains product cards.

    Strategy:
      1. Try fast `requests` (with one retry + a rotating User-Agent).
      2. If the site returns a bot-check / CAPTCHA page, or no product cards,
         fall back to a real headless Chrome browser (Selenium) that passes
         the check.
    Returns None only if both paths fail.
    """
    session = requests.Session()

    for attempt in range(2):
        soup, html = _fetch_requests(url, session)
        if soup is not None:
            if soup.select(card_selector):
                return soup                      # got real results, fast path
            if not looks_blocked(html):
                return soup                      # page loaded, just no matches
            print(f"    [i] {site_name}: bot-check page detected via requests.")
        if attempt == 0:
            time.sleep(random.uniform(2, 4))     # brief backoff before retry

    # Fallback: real browser handles the CAPTCHA / JS challenge.
    print(f"    [i] {site_name}: switching to browser (Selenium) fallback...")
    soup = _fetch_selenium(url, card_selector)
    if soup is not None and soup.select(card_selector):
        print(f"    [+] {site_name}: browser fallback succeeded.")
    elif soup is None:
        print(f"    [!] {site_name}: browser fallback unavailable/failed.")
    return soup


# ------------------------------- scrapers --------------------------------- #

def scrape_flipkart(query, max_items=MAX_ITEMS_PER_SITE):
    """Scrape top search results from Flipkart. Returns list of dicts.

    Flipkart obfuscates and *rotates* its CSS class names, so instead of
    relying on classes we split each product card's visible text into
    segments and identify fields by their shape:
      - Name   -> first long text segment (product titles are lengthy)
      - Rating -> first segment that looks like "4.2"
      - Price  -> first segment starting with the ₹ symbol
    """
    print(f"[Flipkart] Searching for: {query}")
    url = f"https://www.flipkart.com/search?q={requests.utils.quote(query)}"
    soup = get_page(url, "div[data-id]", "Flipkart")
    products = []
    if soup is None:
        print("    [!] Could not load Flipkart search page.")
        return products

    # Product cards carry a data-id attribute; this survives most re-designs.
    cards = soup.select("div[data-id]")
    if not cards:
        print("    [!] No product cards found (layout may have changed).")
        return products

    for card in cards:
        if len(products) >= max_items:
            break
        try:
            segments = [
                s.strip() for s in card.get_text("|", strip=True).split("|") if s.strip()
            ]

            name = next(
                (s for s in segments
                 if len(s) >= 25 and "Add to Compare" not in s and "₹" not in s),
                None,
            )
            rating = next(
                (clean_rating(s) for s in segments if re.fullmatch(r"[0-5]\.\d", s)),
                None,
            )
            # First ₹ segment is the selling price (exchange/strike-through
            # offers appear later in the card).
            price = next(
                (clean_price(s) for s in segments if s.startswith("₹")),
                None,
            )

            if name and price:  # skip ad tiles / accessories without both
                products.append(
                    {"Website": "Flipkart", "Name": name, "Price (INR)": price, "Rating": rating}
                )
        except Exception as exc:  # never let one bad card stop the run
            print(f"    [!] Skipped one card: {exc}")

    print(f"    [+] Scraped {len(products)} products from Flipkart.")
    return products


def scrape_amazon(query, max_items=MAX_ITEMS_PER_SITE):
    """Scrape top search results from Amazon.in. Returns list of dicts."""
    print(f"[Amazon]   Searching for: {query}")
    url = f"https://www.amazon.in/s?k={requests.utils.quote(query)}"
    soup = get_page(url, 'div[data-component-type="s-search-result"]', "Amazon")
    products = []
    if soup is None:
        print("    [!] Could not load Amazon search page.")
        return products

    cards = soup.select('div[data-component-type="s-search-result"]')
    if not cards:
        print("    [!] No product cards found (possible bot check page).")
        return products

    for card in cards:
        if len(products) >= max_items:
            break
        try:
            # A card can hold two <h2>s: a short brand badge ("HP") and the
            # real product title. Take the longest h2 text to get the title.
            h2_texts = [h2.get_text(strip=True) for h2 in card.find_all("h2")]
            name = max(h2_texts, key=len) if h2_texts else None

            price_el = card.select_one("span.a-price > span.a-offscreen")
            price = clean_price(price_el.get_text(strip=True)) if price_el else None

            rating_el = card.select_one("span.a-icon-alt")
            rating = clean_rating(rating_el.get_text(strip=True)) if rating_el else None

            if name and price:
                products.append(
                    {"Website": "Amazon", "Name": name, "Price (INR)": price, "Rating": rating}
                )
        except Exception as exc:
            print(f"    [!] Skipped one card: {exc}")

    print(f"    [+] Scraped {len(products)} products from Amazon.")
    return products


# ------------------------------ reporting --------------------------------- #

def _sanitize(name):
    """Make a product name safe to use as a folder / file name."""
    return re.sub(r'[/\\:*?"<>|]', "_", name).strip()


def prepare_product_paths(query):
    """Resolve Products/<ProductName>/ for this search and return the output
    paths (csv_path, xlsx_path, images_dir).

    Creates the folder if missing. If it already exists, its previous report
    files are removed (old .xlsx/.csv) and images/ is cleared, so the folder
    reflects only the latest search - never a mix of old and new files. Only
    the output *destination* is decided here; report content is untouched.
    """
    safe = _sanitize(query)
    product_dir = os.path.join(PRODUCTS_DIR, safe)
    images_dir = os.path.join(product_dir, IMAGES_SUBDIR)
    os.makedirs(product_dir, exist_ok=True)

    # Replace previous contents of this one product folder.
    for entry in os.listdir(product_dir):
        full = os.path.join(product_dir, entry)
        if os.path.isfile(full) and entry.lower().endswith((".xlsx", ".csv")):
            try:
                os.remove(full)
            except OSError:
                pass  # e.g. open in Excel; build_excel falls back to a new name
    shutil.rmtree(images_dir, ignore_errors=True)
    os.makedirs(images_dir, exist_ok=True)

    base = f"{safe}_{time.strftime('%Y%m%d')}"
    csv_path = os.path.join(product_dir, base + ".csv")
    xlsx_path = os.path.join(product_dir, base + ".xlsx")
    return product_dir, csv_path, xlsx_path, images_dir


def _summary_table(df):
    """Per-website price statistics used by the report and the console."""
    return (
        df.groupby("Website")["Price (INR)"]
        .agg(Count="count", Min="min", Mean="mean", Median="median", Max="max")
        .round(0)
        .astype("Int64")
    )


def _difference_table(df):
    """Head-to-head price differences between the two websites.

    Returns a DataFrame with one row per statistic (cheapest / average /
    most expensive) and the gap between the sites, so the price *difference*
    is stated explicitly (guideline: 'plot basic price difference chart').
    """
    stats = df.groupby("Website")["Price (INR)"].agg(["min", "mean", "max"]).round(0)
    sites = list(stats.index)
    rows = []
    for metric, label in [("min", "Cheapest"), ("mean", "Average"), ("max", "Most expensive")]:
        row = {"Metric": label}
        for site in sites:
            row[site] = int(stats.loc[site, metric])
        if len(sites) == 2:
            gap = int(stats.loc[sites[0], metric] - stats.loc[sites[1], metric])
            row[f"Difference ({sites[0]} - {sites[1]})"] = gap
        rows.append(row)
    return pd.DataFrame(rows)


def build_report(products, query, csv_path):
    """Build the DataFrame and save the CSV report; return the DataFrame."""
    df = pd.DataFrame(products, columns=["Website", "Name", "Price (INR)", "Rating"])
    df.sort_values(["Website", "Price (INR)"], inplace=True, ignore_index=True)

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"[+] Saved: {csv_path}")

    print("\n=== Price Summary by Website ===")
    print(_summary_table(df))
    return df


# ------------------------------- charts ----------------------------------- #

def _short(name, n=32):
    return f"{name[:n]}..." if len(name) > n else name


def chart_prices_by_product(df, query, outdir):
    """Bar chart: every product's price, colored by website."""
    fig, ax = plt.subplots(figsize=(14, 7))
    bar_colors = [COLORS.get(s, "gray") for s in df["Website"]]
    bars = ax.bar(range(len(df)), df["Price (INR)"], color=bar_colors)
    ax.set_xticks(range(len(df)))
    ax.set_xticklabels([_short(n) for n in df["Name"]], rotation=60, ha="right", fontsize=7)
    ax.set_ylabel("Price (INR)")
    ax.set_title(f'Product Prices for "{query}"')
    for bar, price in zip(bars, df["Price (INR)"]):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(),
                f"{price:,}", ha="center", va="bottom", fontsize=6, rotation=90)
    handles = [plt.Rectangle((0, 0), 1, 1, color=c) for c in COLORS.values()]
    ax.legend(handles, COLORS.keys(), title="Website")
    return _save(fig, outdir, "01_prices_by_product.png")


def chart_average_by_site(df, outdir):
    """Bar chart: average price per website (the headline comparison)."""
    avg = df.groupby("Website")["Price (INR)"].mean()
    fig, ax = plt.subplots(figsize=(7, 6))
    bars = ax.bar(avg.index, avg.values, color=[COLORS.get(s, "gray") for s in avg.index])
    ax.set_ylabel("Average Price (INR)")
    ax.set_title("Average Price by Website")
    for bar, val in zip(bars, avg.values):
        ax.text(bar.get_x() + bar.get_width() / 2, val, f"₹{val:,.0f}",
                ha="center", va="bottom", fontsize=11)
    return _save(fig, outdir, "02_average_price_by_site.png")


def chart_price_difference(df, outdir):
    """Grouped bar chart of Min / Average / Max price per website.

    This is the 'basic price difference chart': it puts the two sites
    side-by-side for each statistic so the gap between them is obvious.
    """
    stats = df.groupby("Website")["Price (INR)"].agg(["min", "mean", "max"])
    sites = list(stats.index)
    metrics, labels = ["min", "mean", "max"], ["Minimum", "Average", "Maximum"]
    x = np.arange(len(metrics))
    width = 0.8 / max(len(sites), 1)

    fig, ax = plt.subplots(figsize=(9, 6))
    for i, site in enumerate(sites):
        vals = [stats.loc[site, m] for m in metrics]
        offset = (i - (len(sites) - 1) / 2) * width
        bars = ax.bar(x + offset, vals, width, label=site, color=COLORS.get(site, "gray"))
        for bar, val in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width() / 2, val, f"₹{val:,.0f}",
                    ha="center", va="bottom", fontsize=8)

    # Annotate the average gap between the two sites, if exactly two.
    if len(sites) == 2:
        gap = abs(stats.loc[sites[0], "mean"] - stats.loc[sites[1], "mean"])
        cheaper = stats["mean"].idxmin()
        ax.text(0.5, 0.97, f"Avg. gap: ₹{gap:,.0f}  ({cheaper} cheaper on average)",
                transform=ax.transAxes, ha="center", va="top", fontsize=10,
                bbox=dict(boxstyle="round", fc="#fff3cd", ec="#f8c200"))

    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.set_ylabel("Price (INR)")
    ax.set_title("Price Difference: " + " vs ".join(sites))
    ax.legend(title="Website")
    return _save(fig, outdir, "03_price_difference.png")


def chart_price_distribution(df, outdir):
    """Box plot showing the spread of prices on each website."""
    sites = sorted(df["Website"].unique())
    data = [df.loc[df["Website"] == s, "Price (INR)"].values for s in sites]
    fig, ax = plt.subplots(figsize=(7, 6))
    bp = ax.boxplot(data, tick_labels=sites, patch_artist=True, medianprops=dict(color="black"))
    for patch, s in zip(bp["boxes"], sites):
        patch.set_facecolor(COLORS.get(s, "gray"))
    ax.set_ylabel("Price (INR)")
    ax.set_title("Price Distribution by Website")
    return _save(fig, outdir, "04_price_distribution.png")


def chart_comparison_table(df, outdir):
    """Render the comparison table itself as an image (visual 'table')."""
    show = df.copy()
    show["Name"] = show["Name"].map(lambda n: _short(n, 50))
    show["Price (INR)"] = show["Price (INR)"].map(lambda v: f"₹{v:,}")
    show["Rating"] = show["Rating"].map(lambda r: "—" if pd.isna(r) else f"{r:.1f}")

    fig, ax = plt.subplots(figsize=(12, 0.42 * len(show) + 1.0))
    ax.axis("off")
    ax.set_title("Price Comparison Table", fontsize=13, fontweight="bold", y=0.99)
    tbl = ax.table(cellText=show.values, colLabels=show.columns,
                   cellLoc="left", colLoc="left", bbox=[0, 0, 1, 0.95])
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8)
    # Explicit column proportions so the long Name column gets the room.
    col_widths = [0.11, 0.61, 0.16, 0.12]
    for (row, col), cell in tbl.get_celld().items():
        cell.set_width(col_widths[col])
        if row == 0:  # header
            cell.set_facecolor("#37474f")
            cell.set_text_props(color="white", weight="bold")
        else:         # tint each data row by its website color
            site = show.iloc[row - 1]["Website"]
            cell.set_facecolor("#fdf3c7" if site == "Flipkart" else "#e8eaed")
    return _save(fig, outdir, "05_comparison_table.png")


def _save(fig, outdir, filename):
    """tight_layout + save + close; returns the saved path."""
    path = os.path.join(outdir, filename)
    fig.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"    [+] {path}")
    return path


def make_all_charts(df, query, images_dir):
    """Generate every chart into `images_dir`; return list of paths."""
    os.makedirs(images_dir, exist_ok=True)
    print(f"\n[+] Writing charts to '{images_dir}/':")
    paths = []
    for fn in (
        lambda: chart_prices_by_product(df, query, images_dir),
        lambda: chart_average_by_site(df, images_dir),
        lambda: chart_price_difference(df, images_dir),
        lambda: chart_price_distribution(df, images_dir),
        lambda: chart_comparison_table(df, images_dir),
    ):
        try:
            paths.append(fn())
        except Exception as exc:  # one bad chart shouldn't stop the rest
            print(f"    [!] Chart skipped: {exc}")
    return paths


# ------------------------------- excel ------------------------------------ #

def _write_workbook(path, df, chart_paths, query):
    """Write one .xlsx workbook to `path` (tables + embedded charts)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    summary = _summary_table(df)
    diff = _difference_table(df)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Products", index=False)
        summary.to_excel(writer, sheet_name="Summary")
        diff.to_excel(writer, sheet_name="Price Difference", index=False)
        wb = writer.book

        header_fill = PatternFill("solid", fgColor="37474F")
        header_font = Font(bold=True, color="FFFFFF")

        # --- style the three data sheets: bold header, widths, freeze ---
        for sheet_name in ("Products", "Summary", "Price Difference"):
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            for col_cells in ws.columns:
                col = col_cells[0].column_letter
                longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
                ws.column_dimensions[col].width = min(max(longest + 2, 10), 60)

        # Currency formatting on the Products price column.
        ws_products = writer.sheets["Products"]
        price_idx = list(df.columns).index("Price (INR)") + 1
        price_col = get_column_letter(price_idx)
        for cell in ws_products[price_col][1:]:
            cell.number_format = "₹#,##0"
        ws_products.auto_filter.ref = ws_products.dimensions

        # --- Charts sheet: embed every PNG from Images/ ---
        ws_charts = wb.create_sheet("Charts")
        ws_charts["A1"] = f'Price comparison charts for "{query}"'
        ws_charts["A1"].font = Font(bold=True, size=14)
        row = 3
        for chart_path in chart_paths:
            if not os.path.exists(chart_path):
                continue
            img = XLImage(chart_path)
            # Scale large images down so they fit nicely on the sheet.
            scale = min(1.0, 900 / img.width)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            ws_charts.add_image(img, f"A{row}")
            row += int(img.height / 18) + 3  # ~18px per row + padding


def build_excel(df, chart_paths, query, filename="comparison_report.xlsx"):
    """Save the workbook, falling back to a timestamped name if the target
    file is locked (e.g. currently open in Excel)."""
    try:
        _write_workbook(filename, df, chart_paths, query)
        print(f"[+] Saved: {filename} (Products, Summary, Price Difference, Charts)")
    except PermissionError:
        stem, ext = os.path.splitext(filename)
        alt = f"{stem}_{time.strftime('%H%M%S')}{ext}"  # stay in the same folder
        print(f"[!] '{filename}' is open/locked (close it in Excel). "
              f"Writing to '{alt}' instead.")
        try:
            _write_workbook(alt, df, chart_paths, query)
            print(f"[+] Saved: {alt} (Products, Summary, Price Difference, Charts)")
        except Exception as exc:
            print(f"[!] Excel export failed: {exc}")
    except Exception as exc:
        print(f"[!] Excel export failed: {exc}")


# --------------------------------- main ------------------------------------ #

def parse_args():
    """Command-line interface: an optional product query and per-site count."""
    parser = argparse.ArgumentParser(
        description="Scrape and compare laptop prices across Flipkart & Amazon.in.")
    parser.add_argument(
        "query", nargs="*",
        help="Laptop model to search, e.g. Dell Inspiron 15 (quotes optional). "
             "If omitted, you are prompted for it.")
    parser.add_argument(
        "-n", "--max-items", type=int, default=MAX_ITEMS_PER_SITE,
        help=f"Products to scrape per site (default {MAX_ITEMS_PER_SITE}).")
    return parser.parse_args()


def resolve_query(cli_words):
    """Return the search query from CLI words, or ask for it interactively.

    This is what makes the product dynamic: run the script and simply type a
    different model (or pass it as an argument) - no need to edit the code.
    """
    if cli_words:
        return " ".join(cli_words).strip()
    try:
        entered = input(f"Enter the laptop model to compare "
                        f"[{DEFAULT_QUERY}]: ").strip()
    except EOFError:
        entered = ""  # non-interactive run (e.g. piped) -> use the default
    return entered or DEFAULT_QUERY


def main():
    args = parse_args()
    query = resolve_query(args.query)

    # Respect site usage policy: never let the run exceed MAX_TOTAL_ITEMS.
    per_site = max(1, min(args.max_items, MAX_TOTAL_ITEMS // 2))
    if per_site != args.max_items:
        print(f"[i] Capping to {per_site} items/site to stay under "
              f"{MAX_TOTAL_ITEMS} total (site usage policy).")

    print(f"=== Price Comparison Tool ===\nQuery: {query}\n")

    try:
        all_products = scrape_flipkart(query, per_site)
        polite_sleep()
        all_products += scrape_amazon(query, per_site)
    finally:
        close_driver()  # always release the browser, even on error

    if not all_products:
        print("\n[!] No products scraped from any site. "
              "Check your internet connection or try again later "
              "(sites may be serving a bot-check page).")
        sys.exit(1)

    # Resolve (and reset) this product's dedicated folder under Products/.
    product_dir, csv_path, xlsx_path, images_dir = prepare_product_paths(query)

    df = build_report(all_products, query, csv_path)
    chart_paths = make_all_charts(df, query, images_dir)
    build_excel(df, chart_paths, query, xlsx_path)

    print(f"\n[+] Done. {len(df)} products compared.")
    print(f"    Output folder: {product_dir}")
    print(f"    - {os.path.basename(csv_path)}   (comparison table)")
    print(f"    - {os.path.basename(xlsx_path)}  (tables + embedded charts)")
    print(f"    - {IMAGES_SUBDIR}/  ({len(chart_paths)} chart images)")


if __name__ == "__main__":
    main()
